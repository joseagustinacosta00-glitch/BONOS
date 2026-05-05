from __future__ import annotations

import asyncio
import csv
import io
import re
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Annotated

from fastapi import FastAPI, File, Form, HTTPException, Query, UploadFile, WebSocket, WebSocketDisconnect
from pydantic import BaseModel, Field
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

from backend.ai_assistant import answer_ai_question
from backend.ai_tools import (
    build_basic_study_summary,
    calculate_ratio_points,
    calculate_rolling_zscore,
    calculate_spread_points,
    normalize_ai_ticker,
    summarize_series,
)
from backend.bcra_client import BCRA_SERIES, BcraClient
from backend.bond_calculators import (
    HD_CONVENTION_LABELS,
    LECAP_TICKERS,
    BondHdConvention,
    BondHdCouponInput,
    BondHdFrequency,
    BondHdType,
    BondModelType,
    _days_30_360,
    build_bond_draft,
    build_bond_hd_calculation,
    build_lecap_calculation,
    build_lecap_market_row,
    generate_bond_hd_default_dates,
)
from backend.bonds import BOND_TICKERS
from backend.config import get_settings
from backend.hard_dollar import calculate_hard_dollar_ytm, hard_dollar_bond
from backend.market_calendar import market_calendar, parse_date
from backend.market_data import MarketDataService
from backend.market_scheduler import MarketHistoryScheduler
from backend.market_storage import MarketHistoryStorage
from backend.storage import HISTORICAL_METRIC_TYPES, CalculatorStorage
from backend.time_utils import now_argentina_iso
from backend.time_utils import now_argentina


ROOT_DIR = Path(__file__).resolve().parents[1]
FRONTEND_DIR = ROOT_DIR / "frontend"
PRICE_MARKET_TYPES = ("pesos", "cable", "mep")
SETTLEMENT_TYPES = ("t0", "t1")

settings = get_settings()
market = MarketDataService(settings)
bcra = BcraClient(settings)
storage = CalculatorStorage(ROOT_DIR / settings.app_db_path)

# Backup auto rate-limited: minimo 5 min entre backups despues de un write
import time as _time_module
_last_auto_backup_ts: dict[str, float] = {"ts": 0.0}


def _maybe_backup_after_write(min_interval_seconds: int = 300) -> None:
    """Llama auto_backup() solo si paso suficiente tiempo desde el ultimo.
    Asi cada save HD/LECAP/historical genera un punto de restore reciente
    sin saturar el disco con backups en operaciones rapidas."""
    now = _time_module.time()
    if now - _last_auto_backup_ts["ts"] < min_interval_seconds:
        return
    try:
        storage.auto_backup(retention=30)
        _last_auto_backup_ts["ts"] = now
    except Exception:
        import logging
        logging.getLogger(__name__).exception("backup post-write fallo")
market_history_storage = MarketHistoryStorage(settings.market_history_database_url)
market_scheduler = MarketHistoryScheduler(
    market=market,
    storage=market_history_storage,
    market_open=settings.market_open_local,
    market_close=settings.market_close_local,
    batch_seconds=settings.market_history_batch_seconds,
)

app = FastAPI(title="Monitor de Bonos", version="0.1.0")
app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")


class BondDraftRequest(BaseModel):
    model_type: BondModelType
    issue_date: date
    maturity_date: date
    face_value: float = Field(gt=0)


class LecapCalculationRequest(BaseModel):
    ticker: str
    issue_date: date
    maturity_date: date
    face_value: float = Field(gt=0)
    tem_emission_percent: float


class BondHdCouponPayload(BaseModel):
    payment_date: date
    annual_rate_percent: float = 0.0
    amortization_percent: float = 0.0


class BondHdCalculationRequest(BaseModel):
    issue_date: date
    maturity_date: date
    face_value: float = Field(gt=0)
    bond_type: str = Field(pattern="^(bullet|amortizable|zero_coupon)$")
    frequency: str = Field(pattern="^(annual|semiannual|quarterly|monthly)$")
    convention: str = Field(pattern="^(30_360_eu|30_360_us|180_360_eu|180_360_us|act_360|act_365|act_act)$")
    coupons: list[BondHdCouponPayload] = Field(min_length=1)


class BondHdSavePayload(BaseModel):
    ticker: str = Field(min_length=1, max_length=20)
    issue_date: date
    maturity_date: date
    face_value: float = Field(gt=0)
    bond_type: str = Field(pattern="^(bullet|amortizable|zero_coupon)$")
    frequency: str = Field(pattern="^(annual|semiannual|quarterly|monthly)$")
    convention: str = Field(pattern="^(30_360_eu|30_360_us|180_360_eu|180_360_us|act_360|act_365|act_act)$")
    payload: dict


class BondHdScheduleRequest(BaseModel):
    issue_date: date
    maturity_date: date
    frequency: str = Field(pattern="^(annual|semiannual|quarterly|monthly)$")


class TPlusConversionRequest(BaseModel):
    direction: str = Field(pattern="^(t0_to_t1|t1_to_t0)$")
    price: float = Field(gt=0)
    rate_percent: float | None = None
    use_auto_rate: bool = True


class HistoricalDataRequest(BaseModel):
    ticker: str = Field(min_length=1, max_length=20)
    metric_type: str = Field(pattern="^(parity|dirty_price|clean_price|ytm|tem|tna|volume)$")
    price_market: str = Field(pattern="^(pesos|cable|mep)$")
    settlement_type: str = Field(pattern="^(t0|t1)$")
    value_date: date
    value: float


class AiMemoryNoteRequest(BaseModel):
    title: str = Field(min_length=1, max_length=160)
    category: str = Field(pattern="^(market_concept|calculation_rule|interpretation_rule|study_template|app_behavior|personal_note)$")
    content: str = Field(min_length=1)
    tags: str | None = None


class AiMemoryNoteUpdateRequest(BaseModel):
    title: str | None = Field(default=None, min_length=1, max_length=160)
    category: str | None = Field(default=None, pattern="^(market_concept|calculation_rule|interpretation_rule|study_template|app_behavior|personal_note)$")
    content: str | None = Field(default=None, min_length=1)
    tags: str | None = None
    is_active: bool | None = None


class AiSeriesRef(BaseModel):
    ticker: str = Field(min_length=1, max_length=20)
    dato: str = Field(pattern="^(parity|dirty_price|clean_price|ytm|tem|tna|volume)$")
    mercado: str = Field(pattern="^(pesos|cable|mep)$")
    liquidacion: str = Field(pattern="^(t0|t1)$")


class AiSpreadRequest(BaseModel):
    series_a: AiSeriesRef
    series_b: AiSeriesRef


class AiZScoreRequest(BaseModel):
    points: list[dict[str, object]]
    window: int = Field(default=20, ge=2, le=500)


class AiStudySpreadZScoreRequest(AiSpreadRequest):
    window: int = Field(default=20, ge=2, le=500)
    mode: str = Field(default="spread", pattern="^(spread|ratio)$")


class AssistantMessageRequest(BaseModel):
    message: str = Field(min_length=1)


@app.on_event("startup")
async def startup() -> None:
    import logging
    log = logging.getLogger(__name__)

    db_path = storage.db_path
    db_existed_before = db_path.exists()
    log.info("startup: APP_DB_PATH = %s | existe = %s", db_path, db_existed_before)

    # Auto-restore: si la DB no existe pero hay backups, restaurar del mas
    # reciente automaticamente. Protege contra el escenario donde el contenedor
    # arranca sin la DB previa (deploy con persistent disk recien re-montado,
    # APP_DB_PATH movido, etc.). Los backups viven junto a la DB en /backups.
    if not db_existed_before:
        backups_dir = db_path.parent / "backups"
        if backups_dir.exists():
            candidates = sorted(backups_dir.glob("user_data_*.db"), reverse=True)
            if candidates:
                latest = candidates[0]
                try:
                    storage.restore_from(latest)
                    log.warning(
                        "startup: DB no existia, RESTAURADA automaticamente desde %s (%d bytes)",
                        latest.name, latest.stat().st_size,
                    )
                except Exception as exc:
                    log.exception("startup: auto-restore fallo: %s", exc)

    storage.initialize()

    # Log de contenido para diagnostico
    try:
        counts = storage.get_table_counts()
        log.info("startup: counts = %s", counts)
        if not db_existed_before and not any(counts.values()):
            log.warning(
                "startup: DB recien creada y SIN backups disponibles. "
                "Verificar que APP_DB_PATH apunte al persistent disk de Render "
                "(ej. /var/data/user_data.db) y que el disk este montado."
            )
    except Exception as exc:
        log.exception("startup: get_table_counts fallo: %s", exc)

    try:
        backup_path = storage.auto_backup(retention=30)
        if backup_path:
            log.info("backup automatico creado: %s", backup_path)
    except Exception as exc:
        log.exception("backup automatico fallo: %s", exc)
    await market.start()
    if settings.market_history_enabled:
        try:
            await market_scheduler.start()
        except Exception as exc:
            import logging
            logging.getLogger(__name__).exception("market scheduler no arranco: %s", exc)


@app.on_event("shutdown")
async def shutdown() -> None:
    if settings.market_history_enabled:
        try:
            await market_scheduler.stop()
        except Exception:
            pass
    await market.stop()


_NO_CACHE_HEADERS = {
    "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
    "Pragma": "no-cache",
    "Expires": "0",
}


@app.get("/")
async def index() -> FileResponse:
    return FileResponse(FRONTEND_DIR / "index.html", headers=_NO_CACHE_HEADERS)


@app.get("/app-bundle.js")
async def app_bundle_js() -> FileResponse:
    return FileResponse(
        FRONTEND_DIR / "app.js",
        media_type="application/javascript",
        headers=_NO_CACHE_HEADERS,
    )


@app.get("/styles-bundle.css")
async def styles_bundle_css() -> FileResponse:
    return FileResponse(
        FRONTEND_DIR / "styles.css",
        media_type="text/css",
        headers=_NO_CACHE_HEADERS,
    )


@app.get("/charts-demo")
async def charts_demo() -> FileResponse:
    return FileResponse(FRONTEND_DIR / "charts-demo.html", headers=_NO_CACHE_HEADERS)


@app.get("/ai-demo")
async def ai_demo() -> FileResponse:
    return FileResponse(FRONTEND_DIR / "ai-demo.html", headers=_NO_CACHE_HEADERS)


@app.get("/tradingview")
async def tradingview() -> FileResponse:
    return FileResponse(FRONTEND_DIR / "tradingview.html", headers=_NO_CACHE_HEADERS)


@app.get("/api/tickers")
async def tickers() -> list[dict[str, str]]:
    return [ticker.__dict__ for ticker in BOND_TICKERS]


@app.get("/api/quotes")
async def quotes() -> dict:
    return market.snapshot()


@app.get("/api/bonds/{family}/cashflows")
async def bond_cashflows(family: str, settlement_date: date | None = None) -> dict:
    bond = hard_dollar_bond(family)
    if bond is None:
        raise HTTPException(status_code=404, detail="Bono no soportado.")
    target_date = settlement_date or now_argentina().date()
    return {
        "settlement_date": target_date.isoformat(),
        "bond": bond.to_dict(target_date),
    }


@app.get("/api/bonds/{family}/ytm")
async def bond_ytm(family: str, price: Annotated[float, Query(gt=0)]) -> dict:
    settlement_date = now_argentina().date()
    ytm = calculate_hard_dollar_ytm(family, price, settlement_date)
    if ytm is None:
        raise HTTPException(status_code=422, detail="No se pudo calcular TIR para ese bono/precio.")
    return {
        "family": family.upper().strip(),
        "settlement_date": settlement_date.isoformat(),
        "price": price,
        "ytm": ytm,
    }


@app.get("/api/market/lecaps")
async def market_lecaps(settlement: str = "t1") -> dict:
    settlement_type = _normalize_lecap_settlement(settlement)
    today = now_argentina().date()
    settlement_date = (
        today
        if settlement_type == "t0"
        else market_calendar.next_business_day(today, include_current=False)
    )
    quote_by_ticker = {
        quote["symbol"]: quote for quote in market.lecap_quotes(settlement_type)
    }

    rows = []
    saved_by_ticker = {saved.ticker: saved for saved in storage.list_lecaps()}
    for ticker in LECAP_TICKERS:
        saved = saved_by_ticker.get(ticker)
        quote = quote_by_ticker.get(ticker, {})
        if saved is None:
            rows.append(
                {
                    "ticker": ticker,
                    "settlement_type": settlement_type.upper().replace("T", "T+"),
                    "settlement_date": settlement_date.isoformat(),
                    "maturity_date": None,
                    "effective_payment_date": None,
                    "final_value": None,
                    "days_to_payment": None,
                    "bid": _coerce_optional_float(quote.get("bid")),
                    "offer": _coerce_optional_float(quote.get("ask")),
                    "last": _coerce_optional_float(quote.get("last")),
                    "tna_bid": None,
                    "tna_offer": None,
                    "tna_last": None,
                    "tir_last": None,
                    "tem_last": None,
                    "duration": None,
                    "modified_duration": None,
                    "convexity": None,
                    "updated_at": quote.get("updated_at"),
                }
            )
            continue

        calculation = build_lecap_calculation(
            ticker=saved.ticker,
            issue_date=saved.issue_date,
            maturity_date=saved.maturity_date,
            face_value=saved.face_value,
            tem_emission_percent=saved.tem_emission_percent,
            calendar=market_calendar,
            today=today,
        )
        row = build_lecap_market_row(
            calculation=calculation,
            settlement_type=settlement_type.upper().replace("T", "T+"),
            settlement_date=settlement_date,
            bid=_coerce_optional_float(quote.get("bid")),
            offer=_coerce_optional_float(quote.get("ask")),
            last=_coerce_optional_float(quote.get("last")),
            updated_at=quote.get("updated_at"),
        )
        rows.append(row.to_dict())

    return {
        "status": market.status,
        "source": market.settings.market_source,
        "settlement_type": settlement_type,
        "settlement_date": settlement_date.isoformat(),
        "updated_at": now_argentina_iso(),
        "items": rows,
    }


@app.get("/api/market/caucion/shortest")
async def market_shortest_caucion() -> dict:
    quote = market.shortest_caucion_rate()
    return {
        "status": market.status,
        "source": market.settings.market_source,
        "updated_at": now_argentina_iso(),
        "quote": quote,
    }


@app.get("/api/futures")
async def market_futures() -> dict:
    items = market.futures_quotes()
    return {
        "status": market.status,
        "source": market.settings.market_source,
        "updated_at": now_argentina_iso(),
        "items": items,
    }


@app.get("/api/fx/ratios")
async def fx_ratios() -> dict:
    """Calcula ratios MEP/CCL desde el snapshot actual (precio ARS / precio USD o Cable)."""
    snapshot = market.snapshot()
    by_symbol = {q.get("symbol"): q for q in snapshot.get("quotes", [])}

    pairs = [
        ("AL30", "AL30D", "MEP"),
        ("AL30", "AL30C", "CCL"),
    ]

    items = []
    for ars_symbol, fx_symbol, label in pairs:
        ars_quote = by_symbol.get(ars_symbol)
        fx_quote = by_symbol.get(fx_symbol)
        if not ars_quote or not fx_quote:
            continue
        ars_last = ars_quote.get("last")
        fx_last = fx_quote.get("last")
        ratio = None
        if ars_last is not None and fx_last not in (None, 0):
            try:
                ratio = float(ars_last) / float(fx_last)
            except (TypeError, ValueError, ZeroDivisionError):
                ratio = None
        items.append({
            "name": f"{ars_symbol}/{fx_symbol}",
            "label": label,
            "ars_symbol": ars_symbol,
            "ars_last": ars_last,
            "fx_symbol": fx_symbol,
            "fx_last": fx_last,
            "ratio": ratio,
            "updated_at": ars_quote.get("updated_at") or fx_quote.get("updated_at"),
        })
    return {
        "source": market.settings.market_source,
        "status": market.status,
        "updated_at": now_argentina_iso(),
        "items": items,
    }


@app.get("/api/market/cauciones")
async def market_cauciones() -> dict:
    items = market.caucion_quotes()
    return {
        "status": market.status,
        "source": market.settings.market_source,
        "updated_at": now_argentina_iso(),
        "items": items,
        "shortest": items[0] if items else {},
    }


@app.post("/api/tools/tplus-conversion")
async def tplus_conversion(payload: TPlusConversionRequest) -> dict:
    today = now_argentina().date()
    next_business_day = market_calendar.next_business_day(today, include_current=False)
    calendar_days = (next_business_day - today).days
    caucion_quote = market.shortest_caucion_rate()

    rate_percent = (
        _coerce_optional_float(caucion_quote.get("last"))
        if payload.use_auto_rate
        else payload.rate_percent
    )
    if rate_percent is None:
        raise HTTPException(status_code=422, detail="No hay tasa de caucion disponible.")

    rate_decimal = rate_percent / 100
    factor = 1 + rate_decimal * calendar_days / 365
    if factor <= 0:
        raise HTTPException(status_code=422, detail="La tasa genera un factor invalido.")

    converted_price = (
        payload.price * factor
        if payload.direction == "t0_to_t1"
        else payload.price / factor
    )

    return {
        "direction": payload.direction,
        "today": today.isoformat(),
        "next_business_day": next_business_day.isoformat(),
        "calendar_days": calendar_days,
        "rate_percent": rate_percent,
        "rate_source": "auto" if payload.use_auto_rate else "manual",
        "caucion": caucion_quote,
        "input_price": payload.price,
        "factor": factor,
        "converted_price": converted_price,
    }


@app.get("/api/calendar/summary")
async def calendar_summary() -> dict:
    return market_calendar.summary()


@app.get("/api/calendar/day/{value}")
async def calendar_day(value: str) -> dict:
    target = parse_date(value)
    return {
        "date": target.isoformat(),
        "is_holiday": market_calendar.is_holiday(target),
        "is_business_day": market_calendar.is_business_day(target),
        "previous_business_day": market_calendar.previous_business_day(target).isoformat(),
        "next_business_day": market_calendar.next_business_day(target).isoformat(),
    }


@app.get("/api/calendar/business-days")
async def calendar_business_days(start: str, end: str) -> dict:
    days = market_calendar.business_days_between(parse_date(start), parse_date(end))
    return {"count": len(days), "dates": [item.isoformat() for item in days]}


@app.get("/api/bcra/series")
async def bcra_series(
    desde: date | None = None,
    hasta: date | None = None,
    limit: Annotated[int, Query(ge=0, le=3000)] = 500,
    refresh: bool = False,
) -> dict:
    return bcra.get_all_series(desde=desde, hasta=hasta, force_refresh=refresh, limit=limit)


@app.get("/api/bcra/series/{key}")
async def bcra_one_series(
    key: str,
    desde: date | None = None,
    hasta: date | None = None,
    limit: Annotated[int, Query(ge=0, le=3000)] = 500,
    refresh: bool = False,
) -> dict:
    return bcra._apply_limit(
        bcra.get_series(key, desde=desde, hasta=hasta, force_refresh=refresh),
        limit,
    )


@app.get("/api/bcra/catalog")
async def bcra_catalog() -> dict:
    return {"series": [definition.__dict__ for definition in BCRA_SERIES.values()]}


@app.post("/api/calculators/bond-draft")
async def calculator_bond_draft(payload: BondDraftRequest) -> dict:
    try:
        draft = build_bond_draft(
            model_type=payload.model_type,
            issue_date=payload.issue_date,
            maturity_date=payload.maturity_date,
            face_value=payload.face_value,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return {
        "draft": draft.to_dict(),
        "cashflow_inputs": {
            "issue_date": draft.issue_date.isoformat(),
            "maturity_date": draft.maturity_date.isoformat(),
            "face_value": draft.face_value,
        },
        "next_questions": [],
    }


@app.get("/api/calculators/lecaps/tickers")
async def calculator_lecap_tickers() -> dict:
    return {"tickers": list(LECAP_TICKERS)}


@app.post("/api/calculators/lecaps")
async def calculator_lecaps(payload: LecapCalculationRequest) -> dict:
    try:
        calculation = build_lecap_calculation(
            ticker=payload.ticker,
            issue_date=payload.issue_date,
            maturity_date=payload.maturity_date,
            face_value=payload.face_value,
            tem_emission_percent=payload.tem_emission_percent,
            calendar=market_calendar,
            today=now_argentina().date(),
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return calculation.to_dict()


@app.get("/api/calculators/lecaps/saved")
async def calculator_saved_lecaps() -> dict:
    items = []
    for saved in storage.list_lecaps():
        calculation = build_lecap_calculation(
            ticker=saved.ticker,
            issue_date=saved.issue_date,
            maturity_date=saved.maturity_date,
            face_value=saved.face_value,
            tem_emission_percent=saved.tem_emission_percent,
            calendar=market_calendar,
            today=now_argentina().date(),
        )
        items.append({**saved.to_dict(), "calculation": calculation.to_dict()})
    return {"items": items}


@app.post("/api/calculators/lecaps/saved")
async def calculator_save_lecap(payload: LecapCalculationRequest) -> dict:
    try:
        calculation = build_lecap_calculation(
            ticker=payload.ticker,
            issue_date=payload.issue_date,
            maturity_date=payload.maturity_date,
            face_value=payload.face_value,
            tem_emission_percent=payload.tem_emission_percent,
            calendar=market_calendar,
            today=now_argentina().date(),
        )
        saved = storage.upsert_lecap(
            ticker=payload.ticker,
            issue_date=payload.issue_date,
            maturity_date=payload.maturity_date,
            face_value=payload.face_value,
            tem_emission_percent=payload.tem_emission_percent,
        )
        storage.upsert_cashflows(
            calculator_type="lecap",
            ticker=payload.ticker,
            cashflows=calculation.to_dict()["cashflows"],
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    _maybe_backup_after_write()
    return {"item": saved.to_dict(), "calculation": calculation.to_dict()}


@app.delete("/api/calculators/lecaps/saved/{item_id}")
async def calculator_delete_lecap(item_id: int) -> dict:
    if not storage.delete_lecap(item_id):
        raise HTTPException(status_code=404, detail="LECAP no encontrada.")
    return {"deleted": True}


@app.get("/api/calculators/bond-tamar/calculate")
async def calculator_bond_tamar_calculate(
    issue_date: date,
    maturity_date: date,
    face_value: float = 100.0,
    tem_extra_percent: float = 0.0,
    market_price: float | None = None,
    settlement_type: str = "t1",
) -> dict:
    """Calcula promedio TAMAR sobre la ventana [emision-10BD, vencimiento-10BD],
    TEM TAMAR via formula y VPV proyectado al vencimiento.

    - Valores TAMAR pasados: del BCRA, con carry-forward si falta el dia.
    - Valores TAMAR futuros: promedio simple de los ultimos 5 publicados
      (TAMAR se publica con 2 dias habiles de lag).
    - TAMAR_TEM = [(1 + TAMAR/(365/32))^(365/32)]^(1/12) - 1
      Si tem_extra_percent > 0, se suma a TAMAR_Average antes de aplicar
      la formula (ej. spread sobre TAMAR).
    - VPV = VNO * (1 + TAMAR_TEM)^((DIAS/360) * 12)  con DIAS = vto - emision
    """
    if maturity_date <= issue_date:
        raise HTTPException(status_code=422, detail="Vencimiento debe ser posterior a la emision.")
    if face_value <= 0:
        raise HTTPException(status_code=422, detail="VNO debe ser positivo.")

    series = bcra.get_series("tamar_private_banks_na")
    raw_points = series.get("data", [])
    if not raw_points:
        raise HTTPException(status_code=503, detail="No hay datos TAMAR del BCRA.")

    by_date: dict[date, float] = {}
    for p in raw_points:
        try:
            d = date.fromisoformat(str(p["date"]))
            by_date[d] = float(p["value"])
        except (KeyError, ValueError, TypeError):
            continue
    sorted_dates = sorted(by_date.keys())
    if not sorted_dates:
        raise HTTPException(status_code=503, detail="No hay datos TAMAR validos del BCRA.")
    last_published_date = sorted_dates[-1]

    # 1) TAMAR de emision (igual que tamar-reference)
    issue_minus_10bd = market_calendar.add_business_days(issue_date, -10)
    candidates_emission = [d for d in sorted_dates if d <= issue_minus_10bd]
    tamar_emission = (
        {"reference_date_target": issue_minus_10bd.isoformat(),
         "value_date": candidates_emission[-1].isoformat(),
         "value": by_date[candidates_emission[-1]]}
        if candidates_emission else
        {"reference_date_target": issue_minus_10bd.isoformat(), "value_date": None, "value": None}
    )

    # 2) Proyeccion TAMAR (promedio de los ultimos 5 publicados con lag de 2 BD)
    today = now_argentina().date()
    publication_cutoff = market_calendar.add_business_days(today, -2)
    available = [d for d in sorted_dates if d <= publication_cutoff]
    last5_dates = available[-5:]
    samples = [{"date": d.isoformat(), "value": by_date[d]} for d in last5_dates]
    projection = (sum(by_date[d] for d in last5_dates) / len(last5_dates)) if last5_dates else None
    if projection is None:
        raise HTTPException(status_code=503, detail="Sin datos TAMAR para proyectar.")

    # 3) Promedio TAMAR sobre la ventana [emision-10BD, vencimiento-10BD]
    window_start = issue_minus_10bd
    window_end = market_calendar.add_business_days(maturity_date, -10)
    if window_end <= window_start:
        raise HTTPException(
            status_code=422,
            detail="Plazo del bono insuficiente: ventana TAMAR (emision-10BD a vto-10BD) vacia.",
        )

    # Iterar dias habiles del rango. Para cada uno:
    # - si <= last_published_date: usar TAMAR del dia (carry-forward si falta)
    # - si > last_published_date: usar projection
    daily_values: list[float] = []
    actual_days = 0
    projected_days = 0
    last_known_value = None
    for d in sorted_dates:
        if d > window_start:
            break
        last_known_value = by_date[d]

    cursor = window_start
    while cursor <= window_end:
        if not market_calendar.is_business_day(cursor):
            cursor += timedelta(days=1)
            continue
        if cursor > last_published_date:
            daily_values.append(projection)
            projected_days += 1
        else:
            if cursor in by_date:
                last_known_value = by_date[cursor]
            if last_known_value is not None:
                daily_values.append(last_known_value)
                actual_days += 1
        cursor += timedelta(days=1)

    if not daily_values:
        raise HTTPException(status_code=503, detail="No se pudo construir promedio TAMAR.")
    tamar_average_percent = sum(daily_values) / len(daily_values)

    # 4) TEM TAMAR via formula. TAMAR esta en % NA, convertir a decimal.
    tamar_for_tem_percent = tamar_average_percent + tem_extra_percent
    tamar_decimal = tamar_for_tem_percent / 100.0
    base = 365.0 / 32.0  # ≈ 11.40625
    if 1 + tamar_decimal / base <= 0:
        raise HTTPException(status_code=422, detail="TAMAR resultante invalido para TEM.")
    tea_factor = (1 + tamar_decimal / base) ** base   # 1 + TEA
    tem_factor = tea_factor ** (1.0 / 12.0)            # 1 + TEM
    tamar_tem_decimal = tem_factor - 1.0
    tamar_tem_percent = tamar_tem_decimal * 100.0

    # 5) VPV = VNO * (1 + TEM)^((DIAS/360) * 12)
    # DIAS = conteo 30/360 US (DAYS360 de Excel) entre emision y vencimiento.
    days_total = _days_30_360(issue_date, maturity_date, us=True)
    days_total_calendar = (maturity_date - issue_date).days  # solo informativo
    vpv = face_value * ((1 + tamar_tem_decimal) ** ((days_total / 360.0) * 12.0))

    # 6) Tasa fija TNA contra precio de mercado.
    # settlement_type = "t0" (mismo dia habil) o "t1" (proximo dia habil)
    settlement_offset = 0 if str(settlement_type).lower() == "t0" else 1
    settlement_date = market_calendar.add_business_days(today, settlement_offset)
    days_to_maturity_from_settlement = (maturity_date - settlement_date).days

    fixed_rate_tna_percent = None
    fixed_rate_calc_note = None
    if market_price is not None and market_price > 0:
        if days_to_maturity_from_settlement <= 0:
            fixed_rate_calc_note = "Vencimiento <= liquidacion: no aplica TNA."
        else:
            # tasa_fija = ((VPV / precio_mercado) - 1) / dias * 365
            fixed_rate_decimal = ((vpv / market_price) - 1) / days_to_maturity_from_settlement * 365.0
            fixed_rate_tna_percent = fixed_rate_decimal * 100.0
    elif market_price is None:
        fixed_rate_calc_note = "Cargar precio de mercado para calcular la tasa fija TNA."
    else:
        fixed_rate_calc_note = "Precio de mercado debe ser positivo."

    return {
        "issue_date": issue_date.isoformat(),
        "maturity_date": maturity_date.isoformat(),
        "face_value": face_value,
        "tem_extra_percent": tem_extra_percent,
        "window_start": window_start.isoformat(),
        "window_end": window_end.isoformat(),
        "fixed_rate_warning": {
            "from_date": window_end.isoformat(),
            "to_date": maturity_date.isoformat(),
            "message": (
                f"A partir del {window_end.strftime('%d/%m/%Y')} (vencimiento - 10 dias habiles) "
                f"la tasa pasa a ser fija. Por ahora el calculo no la incluye en el promedio "
                f"y la tasa fija se determinara en la implementacion futura."
            ),
        },
        "tamar_emission": tamar_emission,
        "tamar_maturity_projection": {
            "publication_cutoff": publication_cutoff.isoformat(),
            "today": today.isoformat(),
            "average": projection,
            "samples": samples,
            "sample_count": len(samples),
        },
        "tamar_average_percent": tamar_average_percent,
        "tamar_average_breakdown": {
            "business_days_total": len(daily_values),
            "actual_days": actual_days,
            "projected_days": projected_days,
            "projected_value_used": projection,
            "last_published_date": last_published_date.isoformat(),
        },
        "tamar_for_tem_percent": tamar_for_tem_percent,
        "tamar_tem_percent": tamar_tem_percent,
        "vpv": vpv,
        "vpv_days": days_total,
        "vpv_days_basis": "30/360 US (DAYS360 de Excel)",
        "vpv_days_calendar": days_total_calendar,
        "fixed_rate_calc": {
            "settlement_type": str(settlement_type).lower(),
            "settlement_date": settlement_date.isoformat(),
            "days_to_maturity_from_settlement": days_to_maturity_from_settlement,
            "market_price": market_price,
            "vpv_at_maturity": vpv,
            "fixed_rate_tna_percent": fixed_rate_tna_percent,
            "formula": "((VPV / precio_mercado) - 1) / dias_liq_a_vto * 365",
            "note": fixed_rate_calc_note,
        },
        "source": "BCRA Estadisticas Monetarias v4 - serie TAMAR bancos privados (id=44)",
    }


@app.get("/api/calculators/bond-tamar/tamar-reference")
async def calculator_bond_tamar_reference(
    issue_date: date,
    maturity_date: date,
) -> dict:
    """Devuelve los dos valores TAMAR de referencia para armar el cashflow:
    - tamar_emission: TAMAR vigente 10 dias habiles antes de la emision.
    - tamar_maturity_projection: promedio de los ultimos 5 valores publicados
      (TAMAR se publica con 2 dias habiles de lag), proyectado hasta vencimiento.
    """
    if maturity_date <= issue_date:
        raise HTTPException(status_code=422, detail="Vencimiento debe ser posterior a la emision.")

    # Cargar serie TAMAR completa desde BCRA (cacheada)
    series = bcra.get_series("tamar_private_banks_na")
    raw_points = series.get("data", [])
    if not raw_points:
        raise HTTPException(status_code=503, detail="No hay datos TAMAR disponibles del BCRA.")

    by_date: dict[date, float] = {}
    for p in raw_points:
        try:
            d = date.fromisoformat(str(p["date"]))
            v = float(p["value"])
            by_date[d] = v
        except (KeyError, ValueError, TypeError):
            continue
    sorted_dates = sorted(by_date.keys())
    if not sorted_dates:
        raise HTTPException(status_code=503, detail="No hay datos TAMAR validos del BCRA.")

    # 1) TAMAR de emision: 10 dias habiles antes de issue_date
    issue_minus_10bd = market_calendar.add_business_days(issue_date, -10)
    candidates_emission = [d for d in sorted_dates if d <= issue_minus_10bd]
    tamar_emission = None
    if candidates_emission:
        ref_date = candidates_emission[-1]
        tamar_emission = {
            "reference_date_target": issue_minus_10bd.isoformat(),
            "value_date": ref_date.isoformat(),
            "value": by_date[ref_date],
        }
    else:
        tamar_emission = {
            "reference_date_target": issue_minus_10bd.isoformat(),
            "value_date": None,
            "value": None,
        }

    # 2) TAMAR proyectado: promedio de los ultimos 5 valores publicados.
    #    "Publicados" = con 2 dias habiles de lag desde hoy.
    today = now_argentina().date()
    publication_cutoff = market_calendar.add_business_days(today, -2)
    available = [d for d in sorted_dates if d <= publication_cutoff]
    last5_dates = available[-5:]
    samples = [{"date": d.isoformat(), "value": by_date[d]} for d in last5_dates]
    average = (sum(by_date[d] for d in last5_dates) / len(last5_dates)) if last5_dates else None

    return {
        "issue_date": issue_date.isoformat(),
        "maturity_date": maturity_date.isoformat(),
        "tamar_emission": tamar_emission,
        "tamar_maturity_projection": {
            "publication_cutoff": publication_cutoff.isoformat(),
            "today": today.isoformat(),
            "average": average,
            "samples": samples,
            "sample_count": len(samples),
        },
        "source": "BCRA Estadisticas Monetarias v4 - serie TAMAR bancos privados (id=44)",
    }


@app.get("/api/calculators/bond-hd/conventions")
async def calculator_bond_hd_conventions() -> dict:
    return {
        "conventions": [
            {"key": value.value, "label": HD_CONVENTION_LABELS[value]}
            for value in BondHdConvention
        ],
        "frequencies": [
            {"key": "annual", "label": "Anual", "periods_per_year": 1},
            {"key": "semiannual", "label": "Semestral", "periods_per_year": 2},
            {"key": "quarterly", "label": "Trimestral", "periods_per_year": 4},
            {"key": "monthly", "label": "Mensual", "periods_per_year": 12},
        ],
        "bond_types": [
            {"key": "bullet", "label": "Bullet"},
            {"key": "amortizable", "label": "Amortizable"},
            {"key": "zero_coupon", "label": "Zero-coupon"},
        ],
    }


@app.post("/api/calculators/bond-hd/schedule")
async def calculator_bond_hd_schedule(payload: BondHdScheduleRequest) -> dict:
    try:
        dates = generate_bond_hd_default_dates(
            issue_date=payload.issue_date,
            maturity_date=payload.maturity_date,
            frequency=BondHdFrequency(payload.frequency),
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return {
        "issue_date": payload.issue_date.isoformat(),
        "maturity_date": payload.maturity_date.isoformat(),
        "frequency": payload.frequency,
        "payment_dates": [item.isoformat() for item in dates],
    }


SPANISH_MONTHS: dict[str, int] = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}

_SPANISH_MONTH_PATTERN = re.compile(
    r"(\d{1,2})\s*(?:de\s+)?(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|setiembre|octubre|noviembre|diciembre)(?:\s*(?:de|del)?\s*(\d{4}))?",
    re.IGNORECASE,
)
# Patron permisivo: dia + palabra arbitraria (validamos despues con fuzzy
# match contra los meses para tolerar OCR malo: "iulio" -> "julio", etc.).
_DAY_WORD_PATTERN = re.compile(
    r"(\d{1,2})\s*(?:de\s+)?([A-Za-zÀ-ſ]{3,15})(?:\s*(?:de|del)?\s*(\d{4}))?",
    re.IGNORECASE,
)


def _fuzzy_month_match(word: str) -> int | None:
    """Devuelve el numero de mes si la palabra se parece a un mes (tolerante a
    errores de OCR / hyphenation / caracteres invisibles). None si no hay
    match razonable."""
    word_clean = re.sub(r"[^a-záéíóúñü]", "", word.lower())
    if not word_clean:
        return None
    if word_clean in SPANISH_MONTHS:
        return SPANISH_MONTHS[word_clean]
    import difflib
    candidates = difflib.get_close_matches(word_clean, SPANISH_MONTHS.keys(), n=1, cutoff=0.7)
    if candidates:
        return SPANISH_MONTHS[candidates[0]]
    return None


def _normalize_unicode_text(text: str) -> str:
    """NFKD + strip combining marks. 'días' -> 'dias', 'júlío' -> 'julio'."""
    import unicodedata
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _brute_force_month_dates(
    text: str,
    todos_markers: list[tuple[int, int]],
) -> set[date]:
    """Fallback brute-force: para cada mes en SPANISH_MONTHS, busca todas las
    ocurrencias en el texto (case-insensitive, tolerando 1 error en el medio
    via approximate matching) y para cada match retrocede hasta ~20 chars
    para encontrar el dia (1-2 digitos)."""
    found: set[date] = set()
    text_norm = _normalize_unicode_text(text).lower()
    all_month_names = set(SPANISH_MONTHS.keys())
    for month_name, month_num in SPANISH_MONTHS.items():
        # Variantes con hasta 1 error: substitution o deletion en cualquier pos
        variants = {month_name}
        # Deletions (1 char)
        for i in range(len(month_name)):
            variants.add(month_name[:i] + month_name[i + 1:])
        # Substitutions (cualquier letra por cualquier otra) - solo si len > 4
        if len(month_name) > 4:
            for i in range(len(month_name)):
                for ch in "abcdefghijklmnopqrstuvwxyz":
                    if ch != month_name[i]:
                        variants.add(month_name[:i] + ch + month_name[i + 1:])
        # CRITICO: filtrar variantes que coincidan con otros meses reales.
        # Por ejemplo, "junio" se genera como variante de "julio" (sub l->n)
        # pero es el mes real junio. Sin este filtro, "junio" en el texto se
        # asignaria a julio.
        other_months = all_month_names - {month_name}
        variants = {v for v in variants if v not in other_months and len(v) >= 4}
        for variant in variants:
            for m in re.finditer(rf"\b{re.escape(variant)}\b", text_norm):
                start = m.start()
                # Buscar 1-2 digitos en los ~20 chars anteriores
                window = text_norm[max(0, start - 20):start]
                day_matches = re.findall(r"(\d{1,2})", window)
                if not day_matches:
                    continue
                day = int(day_matches[-1])  # el ultimo (mas cercano al mes)
                if not (1 <= day <= 31):
                    continue
                # Asignar año: explicito en los siguientes ~30 chars o via todos_markers
                tail = text_norm[m.end():m.end() + 30]
                year_match = re.search(r"\b(20\d{2})\b", tail)
                if year_match:
                    year = int(year_match.group(1))
                else:
                    year = next((y for pos, y in todos_markers if pos > start), None)
                    if year is None:
                        continue
                try:
                    found.add(date(year, month_num, day))
                except ValueError:
                    continue
    return found
_TODOS_DE_PATTERN = re.compile(r"todos?\s*(?:de|del)\s*(\d{4})", re.IGNORECASE)
_BROKEN_TODOS_PATTERN = re.compile(r"to[\s\-‐–\.]+dos?", re.IGNORECASE)
_HYPHEN_LINEBREAK_PATTERN = re.compile(r"[-‐–]\s*\n\s*")
_RECURRENT_HINT_PATTERN = re.compile(
    r"(cada\s+a[ñn]o|todos\s+los\s+a[ñn]os|anual\s*mente|recurrente)",
    re.IGNORECASE,
)
_DAY_MONTH_PATTERN = re.compile(r"\b(\d{1,2})[\/\-\.](\d{1,2})\b(?!\s*[\/\-\.]\s*\d)")
_NUMERIC_DATE_PATTERN = re.compile(r"(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})")
_ISO_DATE_PATTERN = re.compile(r"(\d{4})-(\d{1,2})-(\d{1,2})")
# "9 de julio" / "9 enero" sin año explicito (para detectar dia+mes recurrente)
_DAY_SPANISH_MONTH_PATTERN = re.compile(
    r"\b(\d{1,2})\s*(?:de\s+)?(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|setiembre|octubre|noviembre|diciembre)\b",
    re.IGNORECASE,
)
# "primer pago ... en julio del 2021" / "primer pago ... 2021" / "comienza ... 2021" / "empieza ... 2021"
_FIRST_PAYMENT_YEAR_PATTERN = re.compile(
    r"(?:primer(?:o)?\s+(?:pago|cupon|cup[oó]n)|comienza(?:n)?|empieza(?:n)?|inicia(?:n)?|arranca(?:n)?)"
    r"[^.\n]*?(\d{4})",
    re.IGNORECASE,
)
_FIRST_PAYMENT_MONTH_YEAR_PATTERN = re.compile(
    r"(?:primer(?:o)?\s+(?:pago|cupon|cup[oó]n)|comienza(?:n)?|empieza(?:n)?|inicia(?:n)?|arranca(?:n)?)"
    r"[^.\n]*?(?:en|de|del)\s+"
    r"(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|setiembre|octubre|noviembre|diciembre)"
    r"\s*(?:de\s+|del\s+)?(\d{4})",
    re.IGNORECASE,
)


def _detect_first_payment(text: str) -> tuple[int | None, int | None]:
    """Devuelve (year, month) del primer pago si lo encuentra explicitamente.
    month puede ser None si solo se indico el año."""
    m = _FIRST_PAYMENT_MONTH_YEAR_PATTERN.search(text)
    if m:
        month_name = m.group(1).lower()
        year = int(m.group(2))
        month = SPANISH_MONTHS.get(month_name)
        return year, month
    m = _FIRST_PAYMENT_YEAR_PATTERN.search(text)
    if m:
        return int(m.group(1)), None
    return None, None


def _expand_recurring_day_month(
    text: str,
    issue_date: date | None,
    maturity_date: date | None,
) -> set[date]:
    """Detecta patrones tipo '10/07 y 09/01 de cada año' o
    '9 de julio y 9 de enero de cada año, primer pago en julio del 2021'
    y expande las fechas dentro del rango. Necesita maturity_date para saber
    hasta donde extender. issue_date se usa como fallback de inicio si no hay
    'primer pago en YYYY' explicito."""
    if maturity_date is None:
        return set()
    if not _RECURRENT_HINT_PATTERN.search(text):
        return set()

    pairs: set[tuple[int, int]] = set()
    # 1) Pares numericos tipo 10/07
    for match in _DAY_MONTH_PATTERN.finditer(text):
        a, b = int(match.group(1)), int(match.group(2))
        if a > 12 and 1 <= b <= 12:
            day, month = a, b
        elif b > 12 and 1 <= a <= 12:
            day, month = b, a
        elif 1 <= a <= 31 and 1 <= b <= 12:
            day, month = a, b  # default DD/MM (ARG)
        else:
            continue
        if 1 <= day <= 31 and 1 <= month <= 12:
            pairs.add((day, month))

    # 2) Pares en español tipo "9 de julio" / "9 enero"
    for match in _DAY_SPANISH_MONTH_PATTERN.finditer(text):
        day = int(match.group(1))
        month = SPANISH_MONTHS.get(match.group(2).lower())
        if month and 1 <= day <= 31:
            pairs.add((day, month))

    if not pairs:
        return set()

    # Detectar año/mes de primer pago si se indico
    first_year, first_month = _detect_first_payment(text)

    # Decidir año de inicio
    if first_year is not None:
        start_year = first_year
    elif issue_date is not None:
        start_year = issue_date.year
    else:
        start_year = maturity_date.year - 50  # fallback razonable

    # Si tenemos primer mes, filtrar el primer año a fechas >= ese mes
    from calendar import monthrange
    results: set[date] = set()
    for year in range(start_year, maturity_date.year + 1):
        for day, month in pairs:
            if year == start_year and first_month is not None and month < first_month:
                continue
            last_day = monthrange(year, month)[1]
            actual_day = min(day, last_day)
            try:
                candidate = date(year, month, actual_day)
            except ValueError:
                continue
            if candidate > maturity_date:
                continue
            if issue_date is not None and candidate < issue_date and first_year is None:
                continue
            results.add(candidate)
    return results


def _extract_dates_from_text(
    text: str,
    issue_date: date | None = None,
    maturity_date: date | None = None,
) -> list[date]:
    if not text:
        return []
    # Limpieza agresiva pre-regex: remover caracteres invisibles que los PDFs
    # insertan en line breaks (soft hyphen U+00AD, zero-width space U+200B,
    # zero-width joiners U+200C/U+200D, word joiner U+2060, BOM U+FEFF) y
    # que rompen el matching de palabras como "ju<U+00AD>lio".
    invisibles = "­​‌‍⁠﻿"
    cleaned = "".join(ch for ch in text if ch not in invisibles)
    # NFKD + strip combining marks: 'dias' tildados -> sin tildes, robusto
    cleaned = _normalize_unicode_text(cleaned)
    cleaned = _HYPHEN_LINEBREAK_PATTERN.sub("", cleaned)
    cleaned = _BROKEN_TODOS_PATTERN.sub("todos", cleaned)
    # Unir guiones intra-palabra (caso "ju-lio" sin newline cuando el PDF
    # pega rara): letra-letra (sin espacios) -> letraletra. Seguro en
    # contexto financiero donde los meses no llevan guion.
    cleaned = re.sub(r"([A-Za-zÀ-ſ])-([A-Za-zÀ-ſ])", r"\1\2", cleaned)
    # Normalizar TODO whitespace incluyendo NBSP ( ) y otros unicode
    cleaned = re.sub(r"[\s ]+", " ", cleaned)
    normalized = cleaned
    found: set[date] = set()
    # Patrones recurrentes (solo si tenemos rango)
    found.update(_expand_recurring_day_month(normalized, issue_date, maturity_date))

    todos_markers = [(m.start(), int(m.group(1))) for m in _TODOS_DE_PATTERN.finditer(normalized)]

    for match in _SPANISH_MONTH_PATTERN.finditer(normalized):
        day = int(match.group(1))
        month_name = match.group(2).lower()
        explicit_year = match.group(3)
        if explicit_year:
            year = int(explicit_year)
        else:
            year = next((y for pos, y in todos_markers if pos > match.start()), None)
            if year is None:
                continue
        try:
            found.add(date(year, SPANISH_MONTHS[month_name], day))
        except (ValueError, KeyError):
            continue

    # Pasada complementaria con fuzzy match para tolerar OCR malo
    # (ej: "iulio" / "ju lio" / "juloo" -> julio). Solo agrega fechas nuevas
    # que no haya capturado el regex estricto.
    for match in _DAY_WORD_PATTERN.finditer(normalized):
        day = int(match.group(1))
        word = match.group(2)
        explicit_year = match.group(3)
        month = _fuzzy_month_match(word)
        if month is None:
            continue
        if explicit_year:
            year = int(explicit_year)
        else:
            year = next((y for pos, y in todos_markers if pos > match.start()), None)
            if year is None:
                continue
        try:
            found.add(date(year, month, day))
        except ValueError:
            continue

    for match in _ISO_DATE_PATTERN.finditer(normalized):
        try:
            found.add(date(int(match.group(1)), int(match.group(2)), int(match.group(3))))
        except ValueError:
            continue

    iso_spans = [match.span() for match in _ISO_DATE_PATTERN.finditer(normalized)]

    def _inside_iso(start: int, end: int) -> bool:
        return any(s <= start < e or s < end <= e for s, e in iso_spans)

    for match in _NUMERIC_DATE_PATTERN.finditer(normalized):
        if _inside_iso(match.start(), match.end()):
            continue
        a, b, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        if a > 12 and 1 <= b <= 12:
            day, month = a, b
        elif b > 12 and 1 <= a <= 12:
            day, month = b, a
        else:
            day, month = a, b
        try:
            found.add(date(year, month, day))
        except ValueError:
            continue

    # Pasada brute-force final: para cada mes en SPANISH_MONTHS, escanea el
    # texto buscando ocurrencias (incluso con 1 error) y retrocede para hallar
    # el dia. Captura cualquier cosa que el regex estricto y el fuzzy matching
    # de palabras hayan dejado afuera (caso prospectos con OCR raro).
    found.update(_brute_force_month_dates(normalized, todos_markers))

    return sorted(found)


@app.post("/api/calculators/bond-hd/parse-dates")
async def calculator_bond_hd_parse_dates(
    text: Annotated[str | None, Form()] = None,
    file: Annotated[UploadFile | None, File()] = None,
    issue_date: Annotated[date | None, Form()] = None,
    maturity_date: Annotated[date | None, Form()] = None,
) -> dict:
    if not text and (file is None or not file.filename):
        raise HTTPException(status_code=422, detail="Subir archivo o pegar texto con fechas.")

    found: set[date] = set()

    if file is not None and file.filename:
        try:
            rows = await _read_historical_upload(file)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo: {exc}") from exc

        for row in rows:
            try:
                found.add(_parse_historical_date(row))
                continue
            except (ValueError, KeyError):
                pass
            for value in row.values():
                if value is None:
                    continue
                if isinstance(value, datetime):
                    found.add(value.date())
                    continue
                if isinstance(value, date):
                    found.add(value)
                    continue
                for parsed in _extract_dates_from_text(str(value), issue_date, maturity_date):
                    found.add(parsed)

    if text:
        for parsed in _extract_dates_from_text(text, issue_date, maturity_date):
            found.add(parsed)

    if not found:
        raise HTTPException(status_code=422, detail="No se detectaron fechas en el contenido.")

    sorted_dates = sorted(found)
    return {
        "count": len(sorted_dates),
        "dates": [item.isoformat() for item in sorted_dates],
    }


@app.get("/api/calculators/bond-hd/saved")
async def calculator_bond_hd_saved_list() -> dict:
    return {"items": [item.to_dict() for item in storage.list_bond_hd()]}


@app.get("/api/calculators/bond-hd/saved/{ticker}")
async def calculator_bond_hd_saved_one(ticker: str) -> dict:
    saved = storage.get_bond_hd(ticker)
    if saved is None:
        raise HTTPException(status_code=404, detail="Bono HD no encontrado.")
    return {"item": saved.to_dict()}


@app.post("/api/calculators/bond-hd/saved")
async def calculator_bond_hd_saved_upsert(payload: BondHdSavePayload) -> dict:
    import json as _json
    try:
        saved = storage.upsert_bond_hd(
            ticker=payload.ticker,
            issue_date=payload.issue_date,
            maturity_date=payload.maturity_date,
            face_value=payload.face_value,
            bond_type=payload.bond_type,
            frequency=payload.frequency,
            convention=payload.convention,
            payload_json=_json.dumps(payload.payload, ensure_ascii=True),
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    _maybe_backup_after_write()
    return {"item": saved.to_dict()}


@app.delete("/api/calculators/bond-hd/saved/{ticker}")
async def calculator_bond_hd_saved_delete(ticker: str) -> dict:
    if not storage.delete_bond_hd(ticker):
        raise HTTPException(status_code=404, detail="Bono HD no encontrado.")
    return {"deleted": True}


@app.post("/api/calculators/bond-hd")
async def calculator_bond_hd(payload: BondHdCalculationRequest) -> dict:
    try:
        calculation = build_bond_hd_calculation(
            issue_date=payload.issue_date,
            maturity_date=payload.maturity_date,
            face_value=payload.face_value,
            bond_type=BondHdType(payload.bond_type),
            frequency=BondHdFrequency(payload.frequency),
            convention=BondHdConvention(payload.convention),
            coupons=[
                BondHdCouponInput(
                    payment_date=coupon.payment_date,
                    annual_rate_percent=coupon.annual_rate_percent,
                    amortization_percent=coupon.amortization_percent,
                )
                for coupon in payload.coupons
            ],
            calendar=market_calendar,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return calculation.to_dict()


@app.get("/api/calculators/cashflows")
async def calculator_cashflows(calculator_type: str | None = None, ticker: str | None = None) -> dict:
    return {
        "items": [
            cashflow.to_dict()
            for cashflow in storage.list_cashflows(calculator_type=calculator_type, ticker=ticker)
        ]
    }


@app.get("/api/data/tickers")
async def data_tickers() -> dict:
    tickers = sorted({ticker.family for ticker in BOND_TICKERS} | set(LECAP_TICKERS) | set(storage.list_tickers()))
    return {"tickers": tickers}


@app.get("/api/historical-data/series")
async def historical_data_series(ticker: str | None = None) -> dict:
    """Devuelve el catalogo de series cargadas, opcionalmente filtrado por ticker."""
    items = storage.list_historical_series()
    if ticker:
        normalized = _normalize_base_ticker(ticker)
        items = [item for item in items if item.get("ticker") == normalized]
    return {"items": items}


@app.get("/api/historical-data/types")
async def historical_data_types() -> dict:
    return {
        "types": [
            {"key": "parity", "label": "Paridad"},
            {"key": "dirty_price", "label": "Precio dirty"},
            {"key": "clean_price", "label": "Precio clean"},
            {"key": "ytm", "label": "TIR"},
            {"key": "tem", "label": "TEM"},
            {"key": "tna", "label": "TNA"},
            {"key": "volume", "label": "Volumen"},
        ],
        "price_markets": [
            {"key": "pesos", "label": "PESOS"},
            {"key": "cable", "label": "CABLE"},
            {"key": "mep", "label": "MEP"},
        ],
        "settlements": [
            {"key": "t0", "label": "T+0"},
            {"key": "t1", "label": "T+1"},
        ],
        "supported": list(HISTORICAL_METRIC_TYPES),
    }


@app.get("/api/historical-data")
async def historical_data(
    ticker: str | None = None,
    metric_type: str | None = None,
    price_market: str | None = None,
    settlement_type: str | None = None,
    limit: Annotated[int, Query(ge=1, le=5000)] = 500,
) -> dict:
    if metric_type is not None and metric_type not in HISTORICAL_METRIC_TYPES:
        raise HTTPException(status_code=422, detail="Tipo de dato historico no soportado.")
    try:
        normalized_market = _normalize_price_market(price_market) if price_market else None
        normalized_settlement = _normalize_settlement_type(settlement_type) if settlement_type else None
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return {
        "items": [
            point.to_dict()
            for point in storage.list_historical_data(
                ticker=_normalize_base_ticker(ticker) if ticker else None,
                metric_type=metric_type,
                price_market=normalized_market,
                settlement_type=normalized_settlement,
                limit=limit,
            )
        ],
        "series": storage.list_historical_series(),
    }


@app.post("/api/historical-data")
async def save_historical_data(payload: HistoricalDataRequest) -> dict:
    try:
        point, replaced = storage.upsert_historical_data(
            ticker=_normalize_base_ticker(payload.ticker),
            metric_type=payload.metric_type,
            price_market=_normalize_price_market(payload.price_market),
            settlement_type=_normalize_settlement_type(payload.settlement_type),
            value_date=payload.value_date,
            value=payload.value,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    _maybe_backup_after_write()
    return {"item": point.to_dict(), "replaced": replaced}


@app.post("/api/historical-data/upload")
async def upload_historical_data(
    ticker: Annotated[str, Form(min_length=1, max_length=20)],
    metric_type: Annotated[str, Form()],
    price_market: Annotated[str, Form()],
    settlement_type: Annotated[str, Form()],
    file: UploadFile = File(...),
) -> dict:
    rows = await _read_historical_upload(file)
    try:
        normalized_metric = _normalize_metric_type(metric_type)
        normalized_market = _normalize_price_market(price_market)
        normalized_settlement = _normalize_settlement_type(settlement_type)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    imported = 0
    replaced = 0
    errors = []
    detected_format = _detect_date_format(rows)
    for index, row in enumerate(rows, start=2):
        try:
            for row_metric_type, value in _historical_points_from_row(row, normalized_metric):
                _, did_replace = storage.upsert_historical_data(
                    ticker=_normalize_base_ticker(ticker),
                    metric_type=row_metric_type,
                    price_market=_normalize_price_market(
                        _pick_value(row, ("mercado", "market", "price_market")) or normalized_market
                    ),
                    settlement_type=_normalize_settlement_type(
                        _pick_value(row, ("liquidacion", "settlement", "settlement_type")) or normalized_settlement
                    ),
                    value_date=_parse_historical_date(row, date_format=detected_format),
                    value=value,
                )
                imported += 1
                if did_replace:
                    replaced += 1
        except ValueError as exc:
            errors.append({"row": index, "detail": str(exc)})

    if imported == 0:
        detail = errors[:20] if errors else "No se encontraron filas importables en el archivo."
        raise HTTPException(status_code=422, detail=detail)

    if imported or replaced:
        _maybe_backup_after_write()
    return {
        "ticker": _normalize_base_ticker(ticker),
        "metric_type": normalized_metric,
        "price_market": normalized_market,
        "settlement_type": normalized_settlement,
        "imported": imported,
        "replaced": replaced,
        "errors": errors[:20],
        "date_format_detected": detected_format,
    }


@app.get("/api/historical-data/export")
async def export_historical_data(
    ticker: str,
    metric_type: str,
    price_market: str,
    settlement_type: str,
) -> StreamingResponse:
    if metric_type not in HISTORICAL_METRIC_TYPES:
        raise HTTPException(status_code=422, detail="Tipo de dato historico no soportado.")
    try:
        normalized_market = _normalize_price_market(price_market)
        normalized_settlement = _normalize_settlement_type(settlement_type)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    items = storage.list_historical_data(
        ticker=_normalize_base_ticker(ticker),
        metric_type=metric_type,
        price_market=normalized_market,
        settlement_type=normalized_settlement,
        limit=5000,
    )
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Fecha", "Ticker", "Mercado", "Liquidacion", "Dato", "Valor"])
    for item in reversed(items):
        writer.writerow([
            item.value_date.strftime("%d/%m/%Y"),
            item.ticker,
            item.price_market.upper(),
            item.settlement_type.upper().replace("T", "T+"),
            item.metric_type,
            item.value,
        ])
    filename = f"{ticker.upper()}_{metric_type}_{normalized_market}_{normalized_settlement}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/ai/context")
async def ai_context() -> dict:
    return _build_ai_context()


@app.get("/api/ai/memory")
async def ai_memory(
    category: str | None = None,
    search: str | None = None,
    include_inactive: bool = False,
) -> dict:
    return {
        "items": [
            note.to_dict()
            for note in storage.list_ai_memory_notes(
                category=category,
                search=search,
                include_inactive=include_inactive,
            )
        ],
    }


@app.post("/api/ai/memory")
async def create_ai_memory(payload: AiMemoryNoteRequest) -> dict:
    note = storage.create_ai_memory_note(
        title=payload.title,
        category=payload.category,
        content=payload.content,
        tags=payload.tags,
    )
    return {"item": note.to_dict()}


@app.put("/api/ai/memory/{note_id}")
async def update_ai_memory(note_id: int, payload: AiMemoryNoteUpdateRequest) -> dict:
    note = storage.update_ai_memory_note(
        note_id,
        title=payload.title,
        category=payload.category,
        content=payload.content,
        tags=payload.tags,
        is_active=payload.is_active,
    )
    if note is None:
        raise HTTPException(status_code=404, detail="Nota de memoria no encontrada.")
    return {"item": note.to_dict()}


@app.delete("/api/ai/memory/{note_id}")
async def delete_ai_memory(note_id: int) -> dict:
    if not storage.deactivate_ai_memory_note(note_id):
        raise HTTPException(status_code=404, detail="Nota de memoria no encontrada.")
    return {"deleted": True, "mode": "deactivated"}


@app.get("/api/ai/available-series")
async def ai_available_series() -> dict:
    return {
        "items": [
            {
                "ticker": item["ticker"],
                "dato": item["metric_type"],
                "mercado": item["price_market"],
                "liquidacion": item["settlement_type"],
                "first_date": item["first_date"],
                "last_date": item["last_date"],
                "points_count": item["count"],
            }
            for item in storage.list_historical_series()
        ]
    }


@app.get("/api/ai/historical-series")
async def ai_historical_series(
    ticker: str,
    dato: str,
    mercado: str,
    liquidacion: str,
    limit: Annotated[int, Query(ge=1, le=5000)] = 5000,
) -> dict:
    ref = AiSeriesRef(ticker=ticker, dato=dato, mercado=mercado, liquidacion=liquidacion)
    points = _load_ai_series_points(ref, limit=limit)
    return {
        "series": _series_ref_to_dict(ref),
        "summary": summarize_series(points),
        "points": points,
    }


@app.post("/api/ai/spread")
async def ai_spread(payload: AiSpreadRequest) -> dict:
    series_a = _load_ai_series_points(payload.series_a)
    series_b = _load_ai_series_points(payload.series_b)
    points = calculate_spread_points(series_a, series_b)
    return {
        "series_a": _series_ref_to_dict(payload.series_a),
        "series_b": _series_ref_to_dict(payload.series_b),
        "summary": summarize_series(points),
        "points": points,
    }


@app.post("/api/ai/zscore")
async def ai_zscore(payload: AiZScoreRequest) -> dict:
    points = calculate_rolling_zscore(payload.points, payload.window)
    return {"window": payload.window, "summary": summarize_series(points), "points": points}


@app.post("/api/ai/study/spread-zscore")
async def ai_study_spread_zscore(payload: AiStudySpreadZScoreRequest) -> dict:
    series_a = _load_ai_series_points(payload.series_a)
    series_b = _load_ai_series_points(payload.series_b)
    points = (
        calculate_ratio_points(series_a, series_b)
        if payload.mode == "ratio"
        else calculate_spread_points(series_a, series_b)
    )
    zscore_points = calculate_rolling_zscore(points, payload.window)
    return build_basic_study_summary(
        series_a=_series_ref_to_dict(payload.series_a),
        series_b=_series_ref_to_dict(payload.series_b),
        points=points,
        zscore_points=zscore_points,
        window=payload.window,
        mode=payload.mode,
    )


@app.post("/api/assistant/message")
async def assistant_message(payload: AssistantMessageRequest) -> dict:
    memory = [note.to_dict() for note in storage.search_ai_memory_notes(payload.message)]
    return answer_ai_question(
        payload.message,
        memory_notes=memory,
        app_context=_build_ai_context(include_docs=False),
    )


@app.get("/api/system/health")
async def system_health() -> dict:
    """Diagnostico completo: SQLite, backups, market history, entorno."""
    db_info = storage.get_db_info()
    db_path_str = str(db_info.get("absolute_path") or db_info.get("path") or "")
    is_persistent = db_path_str.startswith(("/var/data", "/data", "/mnt/")) or "/var/" in db_path_str
    db_info["is_persistent_path"] = is_persistent

    market_history_info = await market_history_storage.health()
    market_history_info["scheduler_status"] = market_scheduler.status

    environment = {
        "market_source": settings.market_source,
        "rofex_environment": settings.rofex_environment,
        "rofex_user_set": bool(settings.rofex_user),
        "app_db_path": settings.app_db_path,
        "market_history_enabled": settings.market_history_enabled,
        "market_history_database_url_set": bool(settings.market_history_database_url),
        "market_open_local": settings.market_open_local,
        "market_close_local": settings.market_close_local,
        "now_argentina": now_argentina_iso(),
        "is_business_day": market_calendar.is_business_day(now_argentina().date()),
    }

    warnings: list[str] = []
    if not is_persistent:
        warnings.append(
            "APP_DB_PATH no apunta a un disco persistente. "
            "En Render configurar disco en /var/data y APP_DB_PATH=/var/data/user_data.db."
        )
    if not db_info.get("writable"):
        warnings.append("La base SQLite no es escribible.")
    if settings.market_history_enabled and not settings.market_history_database_url:
        warnings.append(
            "Pipeline de market history activo pero falta DATABASE_URL. "
            "Sin Postgres no se persisten ticks."
        )
    if db_info.get("backups", {}).get("count", 0) == 0:
        warnings.append("No hay backups todavia. Apreta 'Crear backup en el server' al menos una vez.")

    return {
        "sqlite": db_info,
        "market_history": market_history_info,
        "environment": environment,
        "warnings": warnings,
    }


@app.get("/api/system/quotes-status")
async def system_quotes_status() -> dict:
    """Diagnostico de cotizaciones: que instrumentos tienen data y cuales no."""
    return market.quotes_status()


@app.get("/api/system/disk-check")
async def system_disk_check() -> dict:
    """Verifica que el filesystem donde vive la base sea escribible y persistente."""
    from datetime import datetime
    target = ROOT_DIR / settings.app_db_path
    parent = target.parent
    result = {
        "path_under_test": str(parent),
        "writable": False,
        "test_file": None,
    }
    try:
        parent.mkdir(parents=True, exist_ok=True)
        marker = parent / f"_persistence_check_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        marker.write_text("ok")
        result["writable"] = True
        result["test_file"] = marker.name
        marker.unlink()
    except OSError as exc:
        result["error"] = str(exc)
    return result


@app.get("/api/data/backup/download")
async def data_backup_download() -> FileResponse:
    """Descarga la base SQLite completa como backup binario."""
    from datetime import datetime
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    tmp_path = ROOT_DIR / "data" / "backups" / f"download_{stamp}.db"
    try:
        storage.backup_to(tmp_path)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    filename = f"user_data_backup_{stamp}.db"
    return FileResponse(
        tmp_path,
        media_type="application/octet-stream",
        filename=filename,
    )


@app.get("/api/data/backup/json")
async def data_backup_json() -> dict:
    """Exporta todas las tablas relevantes como JSON portable."""
    return storage.export_all_json()


@app.get("/api/data/backups")
async def data_backups_list() -> dict:
    """Lista los backups automaticos disponibles en data/backups/."""
    return {"items": storage.list_backups()}


@app.post("/api/data/backup/now")
async def data_backup_now() -> dict:
    """Crea un backup manual ahora mismo."""
    try:
        path = storage.auto_backup(retention=30)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    if path is None:
        raise HTTPException(status_code=404, detail="No hay base para backupear todavia.")
    return {"created": path.name}


@app.post("/api/data/restore-latest")
async def data_restore_latest() -> dict:
    """Restaura la base desde el backup mas reciente disponible. Util si la
    DB se vacio por un deploy con persistent disk mal configurado."""
    backups_dir = storage.db_path.parent / "backups"
    if not backups_dir.exists():
        raise HTTPException(status_code=404, detail="No hay carpeta de backups.")
    candidates = sorted(backups_dir.glob("user_data_*.db"), reverse=True)
    if not candidates:
        raise HTTPException(status_code=404, detail="No hay backups disponibles.")
    latest = candidates[0]
    try:
        storage.restore_from(latest)
    except (FileNotFoundError, ValueError) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    counts_after = storage.get_table_counts()
    return {
        "restored_from": latest.name,
        "size_bytes": latest.stat().st_size,
        "counts_after": counts_after,
    }


@app.post("/api/data/restore")
async def data_restore(file: UploadFile = File(...)) -> dict:
    """Restaura la base desde un .db subido. Hace backup defensivo de la base actual."""
    if not file.filename:
        raise HTTPException(status_code=422, detail="Archivo sin nombre.")
    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".db", ".sqlite", ".sqlite3"}:
        raise HTTPException(status_code=422, detail="Solo se acepta archivo .db/.sqlite.")
    from datetime import datetime
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    incoming = ROOT_DIR / "data" / "backups" / f"incoming_{stamp}.db"
    incoming.parent.mkdir(parents=True, exist_ok=True)
    content = await file.read()
    incoming.write_bytes(content)
    try:
        storage.restore_from(incoming)
    except (FileNotFoundError, ValueError) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    finally:
        try:
            incoming.unlink()
        except OSError:
            pass
    return {"restored": True, "filename": file.filename, "size_bytes": len(content)}


@app.get("/api/market-history/health")
async def market_history_health() -> dict:
    info = await market_history_storage.health()
    info["scheduler_status"] = market_scheduler.status
    return info


@app.get("/api/market-history/instruments")
async def market_history_instruments() -> dict:
    items = await market_history_storage.list_instruments()
    return {"items": [item.to_dict() for item in items]}


@app.get("/api/market-history/ticks")
async def market_history_ticks(
    symbol: str,
    date_from: datetime,
    date_to: datetime,
    interval_seconds: int | None = None,
    limit: Annotated[int, Query(ge=1, le=20000)] = 5000,
) -> dict:
    instrument = await market_history_storage.get_instrument_by_symbol(symbol)
    if instrument is None:
        raise HTTPException(status_code=404, detail="Instrumento no encontrado.")
    if interval_seconds:
        rows = await market_history_storage.fetch_ticks_bucketed(
            instrument_id=instrument.id,
            date_from=date_from,
            date_to=date_to,
            interval_seconds=interval_seconds,
            limit=limit,
        )
    else:
        rows = await market_history_storage.fetch_ticks(
            instrument_id=instrument.id,
            date_from=date_from,
            date_to=date_to,
            limit=limit,
        )
    return {
        "symbol": instrument.symbol,
        "category": instrument.category,
        "interval_seconds": interval_seconds,
        "items": rows,
    }


@app.get("/api/market-history/daily")
async def market_history_daily(
    symbol: str,
    date_from: date,
    date_to: date,
) -> dict:
    instrument = await market_history_storage.get_instrument_by_symbol(symbol)
    if instrument is None:
        raise HTTPException(status_code=404, detail="Instrumento no encontrado.")
    rows = await market_history_storage.fetch_daily_summary(
        instrument_id=instrument.id,
        date_from=date_from,
        date_to=date_to,
    )
    return {
        "symbol": instrument.symbol,
        "category": instrument.category,
        "items": [row.to_dict() for row in rows],
    }


@app.post("/api/market-history/rollup")
async def market_history_rollup(target: date | None = None) -> dict:
    target_date = target or now_argentina().date()
    count = await market_scheduler.force_rollup(target_date)
    return {"date": target_date.isoformat(), "instruments_rolled": count}


@app.get("/api/market-history/export")
async def market_history_export(
    symbol: str,
    date_from: datetime,
    date_to: datetime,
    format: str = Query(default="csv", pattern="^(csv|parquet)$"),
) -> StreamingResponse:
    instrument = await market_history_storage.get_instrument_by_symbol(symbol)
    if instrument is None:
        raise HTTPException(status_code=404, detail="Instrumento no encontrado.")
    rows = await market_history_storage.fetch_ticks(
        instrument_id=instrument.id,
        date_from=date_from,
        date_to=date_to,
        limit=20000,
    )
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["ts", "symbol", "last", "delta_volume", "cumulative_volume"])
        for row in rows:
            writer.writerow([
                row["ts"],
                instrument.symbol,
                row["last"],
                row["delta_volume"],
                row["cumulative_volume"],
            ])
        filename = f"{instrument.symbol}_ticks.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    try:
        import pyarrow as pa
        import pyarrow.parquet as pq
    except ImportError as exc:
        raise HTTPException(status_code=422, detail="pyarrow no instalado.") from exc

    table = pa.table({
        "ts": [row["ts"] for row in rows],
        "symbol": [instrument.symbol for _ in rows],
        "last": [row["last"] for row in rows],
        "delta_volume": [row["delta_volume"] for row in rows],
        "cumulative_volume": [row["cumulative_volume"] for row in rows],
    })
    buffer = io.BytesIO()
    pq.write_table(table, buffer, compression="snappy")
    buffer.seek(0)
    filename = f"{instrument.symbol}_ticks.parquet"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/health")
async def health() -> dict[str, str]:
    return {
        "status": "ok",
        "market": market.status,
        "timezone": "America/Argentina/Buenos_Aires",
        "now": now_argentina_iso(),
    }


@app.websocket("/ws/quotes")
async def websocket_quotes(websocket: WebSocket) -> None:
    await websocket.accept()
    try:
        while True:
            await websocket.send_json(market.snapshot())
            await asyncio.sleep(1)
    except (WebSocketDisconnect, RuntimeError):
        return


def _normalize_lecap_settlement(value: str) -> str:
    normalized = value.lower().replace("+", "").replace(" ", "")
    return "t0" if normalized in {"t0", "0", "ci"} else "t1"


def _coerce_optional_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    return float(value)


async def _read_historical_upload(file: UploadFile) -> list[dict[str, object]]:
    name = (file.filename or "").lower()
    content = await file.read()
    if name.endswith((".xlsx", ".xlsm")):
        return _read_excel_upload(content)
    if name.endswith(".xls"):
        raise HTTPException(status_code=422, detail="Formato .xls no soportado. Guardalo como .xlsx o CSV.")
    return _read_csv_upload(content)


def _read_csv_upload(content: bytes) -> list[dict[str, object]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    sample = text[:2048]
    delimiter_counts = {candidate: sample.count(candidate) for candidate in (";", ",", "\t")}
    delimiter = max(delimiter_counts, key=delimiter_counts.get)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [dict(row) for row in reader]


def _read_excel_upload(content: bytes) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=422, detail="Falta openpyxl para leer Excel.") from exc

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in rows[1:]
        if any(value is not None and value != "" for value in row)
    ]


def _load_ai_series_points(ref: AiSeriesRef, limit: int = 5000) -> list[dict[str, object]]:
    items = storage.list_historical_data(
        ticker=normalize_ai_ticker(ref.ticker),
        metric_type=ref.dato,
        price_market=_normalize_price_market(ref.mercado),
        settlement_type=_normalize_settlement_type(ref.liquidacion),
        limit=limit,
    )
    return [
        {"date": item.value_date.isoformat(), "value": item.value}
        for item in sorted(items, key=lambda item: item.value_date)
    ]


def _series_ref_to_dict(ref: AiSeriesRef) -> dict[str, str]:
    return {
        "ticker": normalize_ai_ticker(ref.ticker),
        "dato": ref.dato,
        "mercado": _normalize_price_market(ref.mercado),
        "liquidacion": _normalize_settlement_type(ref.liquidacion),
    }


def _build_ai_context(include_docs: bool = True) -> dict[str, object]:
    context: dict[str, object] = {
        "project": "bonos",
        "backend": "FastAPI",
        "frontend": "HTML/CSS/JS estatico",
        "storage": "SQLite via APP_DB_PATH",
        "historical_series_key": "ticker base + dato + mercado + liquidacion + fecha",
        "supported_metrics": list(HISTORICAL_METRIC_TYPES),
        "supported_markets": list(PRICE_MARKET_TYPES),
        "supported_settlements": list(SETTLEMENT_TYPES),
        "rules": [
            "No inventar datos de mercado.",
            "Normalizar tickers por familia, por ejemplo AL30D o AL30C a AL30.",
            "No ejecutar SQL libre generado por IA.",
            "No modificar datos sin confirmacion explicita.",
        ],
    }
    if include_docs:
        docs_dir = ROOT_DIR / "docs"
        docs = []
        for name in (
            "PROJECT_CONTEXT.md",
            "DATA_DICTIONARY.md",
            "FINANCIAL_CONCEPTS.md",
            "AI_ASSISTANT_SPEC.md",
            "STUDY_TEMPLATES.md",
            "EXAMPLES.md",
        ):
            path = docs_dir / name
            if path.exists():
                docs.append({"name": name, "content": path.read_text(encoding="utf-8")[:3000]})
        context["docs"] = docs
    return context


def _parse_historical_date(row: dict[str, object], date_format: str = "auto") -> date:
    raw = _pick_value(row, ("fecha", "date", "value_date", "datetime", "date time"))
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    value = str(raw or "").strip()
    for separator in ("/", "-"):
        parts = value.split(separator)
        if len(parts) != 3:
            continue
        first, second, third = parts
        if len(first) == 4:
            return date(int(first), int(second), int(third))
        if len(third) == 4:
            a, b, year = int(first), int(second), int(third)
            if date_format == "mdy":
                if 1 <= a <= 12 and 1 <= b <= 31:
                    return date(year, a, b)
            elif date_format == "dmy":
                if 1 <= b <= 12 and 1 <= a <= 31:
                    return date(year, b, a)
            # auto: heuristica fila por fila
            if a > 12 and b <= 12:
                return date(year, b, a)
            if b > 12 and a <= 12:
                return date(year, a, b)
            return date(year, b, a)
        return date(int(third), int(second), int(first))
    raise ValueError("Fecha invalida. Usa DD/MM/AAAA.")


def _detect_date_format(rows: list[dict[str, object]]) -> str:
    """Inspecciona todas las fechas del archivo y devuelve 'mdy', 'dmy' o 'auto'.
    Si una fila tiene un dia > 12 en la primer posicion -> dmy.
    Si una fila tiene un dia > 12 en la segunda posicion -> mdy.
    Gana el que tenga evidencia clara. Si ambas o ninguna -> auto (ambiguo)."""
    mdy_evidence = 0
    dmy_evidence = 0
    for row in rows:
        raw = _pick_value(row, ("fecha", "date", "value_date", "datetime", "date time"))
        if isinstance(raw, (date, datetime)):
            continue
        value = str(raw or "").strip()
        for separator in ("/", "-"):
            parts = value.split(separator)
            if len(parts) != 3:
                continue
            first, second, third = parts
            if len(first) == 4 or len(third) != 4:
                break
            try:
                a, b = int(first), int(second)
            except ValueError:
                break
            if a > 12 and 1 <= b <= 12:
                dmy_evidence += 1
            elif b > 12 and 1 <= a <= 12:
                mdy_evidence += 1
            break
    if mdy_evidence > 0 and dmy_evidence == 0:
        return "mdy"
    if dmy_evidence > 0 and mdy_evidence == 0:
        return "dmy"
    return "auto"


def _historical_points_from_row(row: dict[str, object], selected_metric: str | None = None) -> list[tuple[str, float]]:
    if selected_metric:
        raw = _pick_value(row, ("valor", "value", *_HISTORICAL_UPLOAD_COLUMNS[selected_metric]))
        if raw is None:
            raw = _pick_first_numeric_value(row, selected_metric)
        if raw is None:
            raise ValueError("No se encontro columna Valor ni la columna del dato seleccionado.")
        return [(selected_metric, _parse_float(raw))]

    metric = _normalize_historical_metric(_pick_value(row, ("tipo", "dato", "metric_type", "metric")))
    if metric:
        return [(metric, _parse_float(_pick_value(row, ("valor", "value"))))]

    points = []
    for metric_type, aliases in _HISTORICAL_UPLOAD_COLUMNS.items():
        raw = _pick_value(row, aliases)
        if raw is not None and str(raw).strip() != "":
            points.append((metric_type, _parse_float(raw)))
    if not points:
        raise ValueError("No se encontro ningun valor historico en la fila.")
    return points


def _pick_value(row: dict[str, object], aliases: tuple[str, ...]) -> object | None:
    normalized = {_normalize_header(key): value for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(_normalize_header(alias))
        if value is not None and str(value).strip() != "":
            return value
    return None


def _pick_first_numeric_value(row: dict[str, object], selected_metric: str) -> object | None:
    ignored = {
        "fecha",
        "date",
        "value date",
        "datetime",
        "date time",
        "tipo",
        "dato",
        "metric type",
        "metric",
        "mercado",
        "market",
        "price market",
        "liquidacion",
        "settlement",
        "settlement type",
    }
    for key, value in row.items():
        normalized_key = _normalize_header(str(key))
        if normalized_key in ignored:
            continue
        if normalized_key in {"volume", "volumen"} and selected_metric != "volume":
            continue
        if selected_metric == "volume" and normalized_key not in {"volume", "volumen"}:
            continue
        if value is None or str(value).strip() == "":
            continue
        try:
            _parse_float(value)
        except ValueError:
            continue
        return value
    return None


def _normalize_historical_metric(value: object | None) -> str | None:
    if value is None:
        return None
    normalized = _normalize_header(str(value))
    alias_map = {
        "paridad": "parity",
        "parity": "parity",
        "precio dirty": "dirty_price",
        "dirty price": "dirty_price",
        "dirty": "dirty_price",
        "precio clean": "clean_price",
        "clean price": "clean_price",
        "clean": "clean_price",
        "tir": "ytm",
        "ytm": "ytm",
        "tem": "tem",
        "tna": "tna",
        "volumen": "volume",
        "volume": "volume",
    }
    return alias_map.get(normalized)


def _normalize_metric_type(value: object | None) -> str:
    metric = _normalize_historical_metric(value)
    if metric in HISTORICAL_METRIC_TYPES:
        return metric
    raise ValueError("Tipo de dato historico no soportado.")


def _normalize_base_ticker(value: str) -> str:
    ticker = value.upper().strip()
    if len(ticker) > 1 and ticker[-1] in {"D", "C"} and any(char.isdigit() for char in ticker[:-1]):
        return ticker[:-1]
    return ticker


def _normalize_price_market(value: object | None) -> str:
    normalized = _normalize_header(str(value or ""))
    if "pesos" in normalized or normalized == "ars":
        return "pesos"
    if "cable" in normalized:
        return "cable"
    if "mep" in normalized:
        return "mep"
    alias_map = {
        "pesos": "pesos",
        "peso": "pesos",
        "ars": "pesos",
        "$": "pesos",
        "cable": "cable",
        "c": "cable",
        "mep": "mep",
        "dolar mep": "mep",
        "usd": "mep",
    }
    result = alias_map.get(normalized)
    if result not in PRICE_MARKET_TYPES:
        raise ValueError("Mercado invalido. Usa PESOS, CABLE o MEP.")
    return result


def _normalize_settlement_type(value: object | None) -> str:
    normalized = str(value or "").lower().replace("+", "").replace(" ", "").strip()
    if normalized in {"t0", "0", "ci"} or "t0" in normalized:
        return "t0"
    if normalized in {"t1", "1", "24hs", "24h"} or "t1" in normalized:
        return "t1"
    raise ValueError("Liquidacion invalida. Usa T+0 o T+1.")


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value.strip().lower())
    ascii_value = "".join(char for char in normalized if not unicodedata.combining(char))
    return ascii_value.replace("_", " ").replace("-", " ")


def _parse_float(value: object) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("%", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    return float(text)


_HISTORICAL_UPLOAD_COLUMNS = {
    "parity": ("paridad", "parity"),
    "dirty_price": ("precio dirty", "dirty price", "dirty"),
    "clean_price": ("precio clean", "clean price", "clean"),
    "ytm": ("tir", "ytm"),
    "tem": ("tem",),
    "tna": ("tna",),
    "volume": ("volumen", "volume"),
}
